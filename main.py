from requests import get
import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from bs4 import BeautifulSoup
from datetime import date


# ciudades para observar
town = ['malaga', 'torremolinos']
# словарь с диапазонами цен за аренду квартир
prises = {600: 699, 700: 799, 800: 899, 900: 1000}

HOST = 'https://www.milanuncios.com/'
URL = 'https://www.milanuncios.com/alquiler-de-pisos-en-malaga-malaga/?fromSearch\
       =1&desde=700&hasta=799&dormd=3&dormh=3&banosd=2&banosh=2'
HEADERS = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0'
}
lista = []

# функция возвращает содержимое страницы url
def get_html(url, params=None):
    r = get(url, headers=HEADERS, params=params)
    return r

# функция выбирает из заданного содержимого страницы url необходимые данные
def get_content(text_of_page):
    soup = BeautifulSoup(text_of_page, 'html.parser')
    items = str(soup.find('span', style="text-align:left;font-weight:400"))
    count = ''
    for i in items[108:134]:
        if i.isdigit():
            count += i
    if count == '':
        count = 0
    return int(count)

# если запрос выполнен верно и искомая страница существует
if get_html(URL).status_code == 200:
    number = 0
    n = 10
    # для каждого из городов списка town
    for i in town:
        # для каждого диапазона цен из словаря prises
        for j, jj in prises.items():
            # извлекаем содержимое заданной страницы url
            html = get_html(f'https://www.milanuncios.com/alquiler-de-pisos-en-{i}-malaga/?fromSearch\
                =1&desde={j}&hasta={jj}&dormd=3&dormh=3&banosd=2&banosh=2')
            # дообавляем в конец списка lista необходимые выбранные данные
            lista.append(get_content(html.text))
            # visualization of the software process
            number += 1
            n -= 1
            numb = {1: 'st', 2: 'nd', 3: 'rd'}
            if number < 4:
                suff = numb[number]
            else:
                suff = 'th'
            print(f'Parsing of {number}-{suff} page'+' '*n+' ...')
    # для каждого из городов списка town
    for i in town:
        # вставляем нужные нам исходные данные в URL при выборе квартир
        html = get_html(f'https://www.milanuncios.com/alquiler-de-pisos-en-{i}-malaga/?fromSearch\
            =1&desde=600&demanda=n&dormd=4&dormh=4&banosd=2')
        # извлекаем текст из заданного URL
        lista.append(get_content(html.text))
        # visualization of the software process
        number += 1
        n -= 1
        print(f'Parsing of {number}-th page'+' '*n+' ...')
else:
    print('Error')
print(lista)

# работа с файлом Estadistica.xlsm
book = openpyxl.load_workbook(filename='Estadistica.xlsm', read_only=False, keep_vba=True)
sheet = book.active
# next_row - строка, которая добавится в конец существующей таблицы (Estadistica.xlsm)
next_row = sheet.max_row + 1
# переменная для форматирования добавленной строки на текущий лист файла Estadistica.xlsm
medium_border = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))

# если дата первой ячейки последней строки активного листа файла
# Estadistica.xlsm отлична от текущей даты
if str(sheet[sheet.max_row][0].value)[0:10] != str(date.today()):
    # добавляем в начало списка lista текущую дату
    lista.insert(0, date.today())
    # добавляем в активный лист файла Estadistica.xlsm новую строку после последней
    sheet.append(lista)
    # задаём форматирование добавленной строки в активном листе файла Estadistica.xlsm
    sheet.row_dimensions[next_row].height = 23.6
    sheet[next_row][0].fill = PatternFill(fill_type='solid', start_color='FFB6FCC5')
    sheet[next_row][9].fill = PatternFill(fill_type='solid', start_color='FFF2F2F2')
    sheet[next_row][10].fill = PatternFill(fill_type='solid', start_color='FFF2F2F2')
    sheet[next_row][0].font = Font(size=10)
    for i in range(1, 12):
        sheet.cell(row=next_row, column=i).border = medium_border
        sheet.cell(row=next_row, column=i).alignment = Alignment(horizontal='center', vertical='center')
    for i in range(1, 11):
        sheet[next_row][i].font = Font(size=18)
    # задаём числовой формат даты (столбец A) как у предыдущей ячейки сверху
    sheet[next_row][0].number_format = sheet[next_row - 1][0].number_format

# сохраняем изменения в файл Estadistica.xlsm
book.save('Estadistica.xlsm')

# открываем файл для визуального просмотра
import os
os.startfile('Estadistica.xlsm')
